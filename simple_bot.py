import os
import logging
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
from openpyxl import load_workbook
import re
import tempfile
from collections import defaultdict

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ТОКЕН БОТА - ЗАМЕНИ НА СВОЙ!
BOT_TOKEN = "7970625516:AAFDZxWKAXrKxdDc99Ghx4kgJwUarSpoqaI"

# Функции парсинга из твоего рабочего скрипта
def find_table_structure(ws):
    headers_positions = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell_value = str(cell.value).strip()
                if "Товары (работы, услуги)" in cell_value:
                    headers_positions['description'] = (cell.row, cell.column)
                elif "Сумма" in cell_value and cell_value != "Сумма с НДС":
                    headers_positions['amount'] = (cell.row, cell.column)
    return headers_positions

def extract_data_from_description(description):
    description_str = str(description)
    
    # Маршрут (все до первой запятой)
    route = description_str.split(',')[0].strip()
    
    # Дата из текста (формат "от 06.09.25")
    date_match = re.search(r'от\s+(\d{2}\.\d{2}\.\d{2})', description_str)
    date_str = date_match.group(1) if date_match else "Дата не найдена"
    
    # Гос. номер - ищем 3 цифры подряд
    plate_match = re.search(r'(\d{3})', description_str)
    car_plate = plate_match.group(1) if plate_match else "Неизвестно"
    
    # Фамилия водителя
    driver_match = re.search(r',\s*([А-Я][а-я]+)\s+[А-Я]\.[А-Я]\.', description_str)
    if driver_match:
        driver_name = driver_match.group(1)
    else:
        alt_driver_match = re.search(r',\s*([А-Я][а-я]+)', description_str)
        driver_name = alt_driver_match.group(1) if alt_driver_match else "Фамилия не найдена"
    
    return route, date_str, car_plate, driver_name

def parse_invoice_file(file_path):
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        headers = find_table_structure(ws)
        
        if not headers.get('description') or not headers.get('amount'):
            return []
        
        header_row = max(h[0] for h in headers.values())
        description_col = headers['description'][1]
        amount_col = headers['amount'][1]
        
        parsed_data = []
        row_num = header_row + 1
        current_empty_rows = 0
        max_empty_rows = 5
        
        while current_empty_rows < max_empty_rows:
            description_cell = ws.cell(row=row_num, column=description_col)
            description = description_cell.value
            
            if not description:
                current_empty_rows += 1
                row_num += 1
                continue
                
            current_empty_rows = 0
            description_str = str(description)
            
            if any(word in description_str.lower() for word in ['итого', 'всего', 'итог', 'сумма']):
                row_num += 1
                continue
            
            amount_cell = ws.cell(row=row_num, column=amount_col)
            amount = amount_cell.value
            
            if amount is not None:
                try:
                    if isinstance(amount, str) and any(char.isalpha() for char in amount.replace(' ', '').replace(',', '.')):
                        row_num += 1
                        continue
                    
                    amount_str = str(amount).replace(' ', '').replace(',', '.')
                    amount_value = float(amount_str)
                    
                    route, date_str, car_plate, driver_name = extract_data_from_description(description_str)
                    
                    if car_plate != "Неизвестно" and amount_value > 0:
                        parsed_data.append({
                            'Дата': date_str,
                            'Маршрут': route,
                            'Стоимость': amount_value,
                            'Гос_номер': car_plate,
                            'Водитель': driver_name
                        })
                    
                except (ValueError, TypeError):
                    pass
            
            row_num += 1
            
            if row_num > header_row + 1000:
                break
        
        return parsed_data
        
    except Exception as e:
        logger.error(f"Ошибка при обработке файла: {e}")
        return []

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    welcome_text = """
🚛 *Transport Analytics Bot*

Отправьте мне Excel файл с транспортными накладными, и я:
• Проанализирую данные по автомобилям
• Покажу итоговые суммы по каждой машине
• Сгенерирую отчет

*Поддерживаемые форматы:* .xlsx, .xls

Просто отправьте файл и получите результат!
    """
    await update.message.reply_text(welcome_text, parse_mode='Markdown')

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик загруженных файлов"""
    try:
        # Получаем информацию о файле
        document = update.message.document
        file = await context.bot.get_file(document.file_id)
        
        # Проверяем что это Excel файл
        if not (document.file_name.endswith('.xlsx') or document.file_name.endswith('.xls')):
            await update.message.reply_text("❌ Пожалуйста, отправьте Excel файл (.xlsx или .xls)")
            return
        
        await update.message.reply_text("🔍 Обрабатываю файл...")
        
        # Скачиваем файл во временную папку
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            await file.download_to_drive(temp_file.name)
            
            # Парсим файл
            data = parse_invoice_file(temp_file.name)
            
            # Удаляем временный файл
            os.unlink(temp_file.name)
        
        if not data:
            await update.message.reply_text("❌ Не удалось найти данные в файле. Проверьте формат.")
            return
        
        # Анализируем данные БЕЗ pandas
        car_stats = defaultdict(lambda: {'total': 0, 'trips': 0, 'drivers': set()})
        
        for item in data:
            car_plate = item['Гос_номер']
            car_stats[car_plate]['total'] += item['Стоимость']
            car_stats[car_plate]['trips'] += 1
            car_stats[car_plate]['drivers'].add(item['Водитель'])
        
        # Формируем отчет по автомобилям
        car_reports = []
        for car_plate, stats in car_stats.items():
            drivers = ', '.join(stats['drivers'])
            car_reports.append(f"🚗 *{car_plate}*\n"
                             f"• Поездок: {stats['trips']}\n"
                             f"• Водители: {drivers}\n"
                             f"• Общая сумма: {stats['total']:,.0f} руб.\n")
        
        # Общая статистика
        total_trips = len(data)
        total_amount = sum(item['Стоимость'] for item in data)
        unique_cars = len(car_stats)
        unique_drivers = len(set(item['Водитель'] for item in data))
        
        # Формируем ответ
        response = f"""
📊 *ОТЧЕТ ПО ФАЙЛУ: {document.file_name}*

*Общая статистика:*
• Всего поездок: {total_trips}
• Автомобилей: {unique_cars}  
• Водителей: {unique_drivers}
• Общая сумма: {total_amount:,.0f} руб.

*По автомобилям:*
{chr(10).join(car_reports)}

✅ Обработка завершена!
        """
        
        await update.message.reply_text(response, parse_mode='Markdown')
        
    except Exception as e:
        logger.error(f"Ошибка: {e}")
        await update.message.reply_text("❌ Произошла ошибка при обработке файла")

async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ошибок"""
    logger.error(f"Ошибка: {context.error}")
    await update.message.reply_text("❌ Произошла ошибка. Попробуйте еще раз.")

def main():
    """Запуск бота"""
    # Создаем приложение
    application = Application.builder().token(BOT_TOKEN).build()
    
    # Добавляем обработчики
    application.add_handler(CommandHandler("start", start))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    application.add_error_handler(error_handler)
    
    # Запускаем бота
    print("🤖 Бот запущен...")
    application.run_polling()

if __name__ == "__main__":
    main()
