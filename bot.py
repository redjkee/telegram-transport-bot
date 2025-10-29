import os
import logging
import asyncio
import threading
from http.server import HTTPServer, BaseHTTPRequestHandler
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import Message
from openpyxl import load_workbook
import re
import tempfile
from collections import defaultdict

# HTTP сервер для Render
class HealthHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"Bot is alive")
    
    def log_message(self, format, *args):
        pass

def run_http_server():
    server = HTTPServer(('0.0.0.0', 8080), HealthHandler)
    server.serve_forever()

# Запускаем HTTP сервер
http_thread = threading.Thread(target=run_http_server, daemon=True)
http_thread.start()

# Остальной код бота
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.getenv('BOT_TOKEN')
user_data_store = defaultdict(list)
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher()

def find_table_structure(ws):
    headers_positions = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell_value = str(cell.value).strip()
                if "Товары (работы, услуги)" in cell_value:
                    headers_positions['description'] = (cell.row, cell.column)
                elif "Сумма" in cell_value and cell_value != "Сумма с НДС":
                    headers_positions['amount'] = (cell.row, cell.column)
    return headers_positions

def extract_data_from_description(description):
    description_str = str(description)
    route = description_str.split(',')[0].strip()
    date_match = re.search(r'от\s+(\d{2}\.\d{2}\.\d{2})', description_str)
    date_str = date_match.group(1) if date_match else "Дата не найдена"
    plate_match = re.search(r'(\d{3})', description_str)
    car_plate = plate_match.group(1) if plate_match else "Неизвестно"
    driver_match = re.search(r',\s*([А-Я][а-я]+)\s+[А-Я]\.[А-Я]\.', description_str)
    if driver_match:
        driver_name = driver_match.group(1)
    else:
        alt_driver_match = re.search(r',\s*([А-Я][а-я]+)', description_str)
        driver_name = alt_driver_match.group(1) if alt_driver_match else "Фамилия не найдена"
    return route, date_str, car_plate, driver_name

def parse_invoice_file(file_path):
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        headers = find_table_structure(ws)
        
        if not headers.get('description') or not headers.get('amount'):
            return []
        
        header_row = max(h[0] for h in headers.values())
        description_col = headers['description'][1]
        amount_col = headers['amount'][1]
        
        parsed_data = []
        row_num = header_row + 1
        current_empty_rows = 0
        max_empty_rows = 5
        
        while current_empty_rows < max_empty_rows:
            description_cell = ws.cell(row=row_num, column=description_col)
            description = description_cell.value
            
            if not description:
                current_empty_rows += 1
                row_num += 1
                continue
                
            current_empty_rows = 0
            description_str = str(description)
            
            if any(word in description_str.lower() for word in ['итого', 'всего', 'итог', 'сумма']):
                row_num += 1
                continue
            
            amount_cell = ws.cell(row=row_num, column=amount_col)
            amount = amount_cell.value
            
            if amount is not None:
                try:
                    if isinstance(amount, str) and any(char.isalpha() for char in amount.replace(' ', '').replace(',', '.')):
                        row_num += 1
                        continue
                    
                    amount_str = str(amount).replace(' ', '').replace(',', '.')
                    amount_value = float(amount_str)
                    
                    route, date_str, car_plate, driver_name = extract_data_from_description(description_str)
                    
                    if car_plate != "Неизвестно" and amount_value > 0:
                        parsed_data.append({
                            'Дата': date_str,
                            'Маршрут': route,
                            'Стоимость': amount_value,
                            'Гос_номер': car_plate,
                            'Водитель': driver_name,
                            'Файл': os.path.basename(file_path)
                        })
                    
                except (ValueError, TypeError):
                    pass
            
            row_num += 1
            
            if row_num > header_row + 1000:
                break
        
        return parsed_data
        
    except Exception as e:
        logger.error(f"Ошибка при обработке файла: {e}")
        return []

def calculate_statistics(data):
    if not data:
        return None
    
    total_trips = len(data)
    total_amount = sum(item['Стоимость'] for item in data)
    
    unique_cars = set(item['Гос_номер'] for item in data)
    unique_drivers = set(item['Водитель'] for item in data)
    unique_files = set(item['Файл'] for item in data)
    
    car_stats = {}
    for item in data:
        car_plate = item['Гос_номер']
        if car_plate not in car_stats:
            car_stats[car_plate] = {
                'total_amount': 0,
                'trips_count': 0,
                'drivers': set(),
                'files': set()
            }
        
        car_stats[car_plate]['total_amount'] += item['Стоимость']
        car_stats[car_plate]['trips_count'] += 1
        car_stats[car_plate]['drivers'].add(item['Водитель'])
        car_stats[car_plate]['files'].add(item['Файл'])
    
    # Статистика по водителям
    driver_stats = {}
    for item in data:
        driver = item['Водитель']
        if driver not in driver_stats:
            driver_stats[driver] = {
                'total_amount': 0,
                'trips_count': 0,
                'cars': set(),
                'files': set()
            }
        driver_stats[driver]['total_amount'] += item['Стоимость']
        driver_stats[driver]['trips_count'] += 1
        driver_stats[driver]['cars'].add(item['Гос_номер'])
        driver_stats[driver]['files'].add(item['Файл'])
    
    return {
        'total_trips': total_trips,
        'total_amount': total_amount,
        'unique_cars': len(unique_cars),
        'unique_drivers': len(unique_drivers),
        'unique_files': len(unique_files),
        'car_stats': car_stats,
        'driver_stats': driver_stats
    }

def calculate_file_statistics(file_data):
    if not file_data:
        return None
    
    total_amount = sum(item['Стоимость'] for item in file_data)
    trips_count = len(file_data)
    
    return {
        'total_amount': total_amount,
        'trips_count': trips_count
    }

@dp.message(Command("start"))
async def start_handler(message: Message):
    welcome_text = """
🚛 Transport Analytics Bot

Отправьте мне Excel файлы с транспортными накладными, и я:
• Соберу данные из ВСЕХ файлов
• Покажу суммарную статистику по автомобилям и водителям
• Сгенерирую общий отчет

📊 ДОСТУПНЫЕ КОМАНДЫ:

/report - полный отчет (авто + водители)
/cars - отчет только по автомобилям  
/drivers - отчет только по водителям
/clear - очистить все данные

🔍 ПОИСК:
Отправьте номер автомобиля (например: 302) или фамилию водителя для получения детальной статистики

📁 Поддерживаемые форматы: .xlsx, .xls

💡 Просто отправляйте файлы один за другим, а затем используйте команды для получения отчетов!
    """
    await message.answer(welcome_text)

@dp.message(Command("clear"))
async def clear_handler(message: Message):
    user_id = message.from_user.id
    user_data_store[user_id] = []
    await message.answer("✅ Все данные очищены! Можно загружать новые файлы.")

@dp.message(Command("report"))
async def report_handler(message: Message):
    user_id = message.from_user.id
    user_data = user_data_store[user_id]
    
    if not user_data:
        await message.answer("📭 Нет данных для отчета. Сначала отправьте файлы.")
        return
    
    await generate_report(message, user_data, "ПОЛНЫЙ ОТЧЕТ")

@dp.message(Command("cars"))
async def cars_handler(message: Message):
    """Отчет только по автомобилям"""
    user_id = message.from_user.id
    user_data = user_data_store[user_id]
    
    if not user_data:
        await message.answer("📭 Нет данных для отчета. Сначала отправьте файлы.")
        return
    
    stats = calculate_statistics(user_data)
    
    car_reports = []
    for car_plate, car_data in stats['car_stats'].items():
        drivers = ', '.join(car_data['drivers'])
        files = ', '.join(list(car_data['files'])[:3])
        if len(car_data['files']) > 3:
            files += f" ... (еще {len(car_data['files']) - 3})"
        
        car_reports.append(f"🚗 {car_plate}\n"
                         f"• Поездок: {car_data['trips_count']}\n"
                         f"• Водители: {drivers}\n"
                         f"• Файлы: {files}\n"
                         f"• Общая сумма: {car_data['total_amount']:,.0f} руб.\n")
    
    response = f"""
📊 ОТЧЕТ ПО АВТОМОБИЛЯМ

Всего автомобилей: {stats['unique_cars']}
Общая сумма: {stats['total_amount']:,.0f} руб.

{chr(10).join(car_reports)}
    """
    
    if len(response) > 4000:
        parts = [response[i:i+4000] for i in range(0, len(response), 4000)]
        for part in parts:
            await message.answer(part)
            await asyncio.sleep(0.5)
    else:
        await message.answer(response)

@dp.message(Command("drivers"))
async def drivers_handler(message: Message):
    """Отчет только по водителям"""
    user_id = message.from_user.id
    user_data = user_data_store[user_id]
    
    if not user_data:
        await message.answer("📭 Нет данных для отчета. Сначала отправьте файлы.")
        return
    
    stats = calculate_statistics(user_data)
    
    driver_reports = []
    for driver, driver_data in stats['driver_stats'].items():
        if driver == "Фамилия не найдена":
            continue
        cars = ', '.join(driver_data['cars'])
        files = ', '.join(list(driver_data['files'])[:3])
        if len(driver_data['files']) > 3:
            files += f" ... (еще {len(driver_data['files']) - 3})"
        
        driver_reports.append(f"👤 {driver}\n"
                            f"• Поездок: {driver_data['trips_count']}\n"
                            f"• Автомобили: {cars}\n"
                            f"• Файлы: {files}\n"
                            f"• Общая сумма: {driver_data['total_amount']:,.0f} руб.\n")
    
    response = f"""
📊 ОТЧЕТ ПО ВОДИТЕЛЯМ

Всего водителей: {len([d for d in stats['driver_stats'].keys() if d != "Фамилия не найдена"])}
Общая сумма: {stats['total_amount']:,.0f} руб.

{chr(10).join(driver_reports)}
    """
    
    if len(response) > 4000:
        parts = [response[i:i+4000] for i in range(0, len(response), 4000)]
        for part in parts:
            await message.answer(part)
            await asyncio.sleep(0.5)
    else:
        await message.answer(response)

@dp.message(F.document)
async def document_handler(message: Message):
    try:
        user_id = message.from_user.id
        document = message.document
        
        if not (document.file_name.endswith('.xlsx') or document.file_name.endswith('.xls')):
            await message.answer("❌ Пожалуйста, отправьте Excel файл (.xlsx или .xls)")
            return
        
        await message.answer(f"🔍 Обрабатываю файл: {document.file_name}")
        
        file = await bot.get_file(document.file_id)
        file_path = f"/tmp/{document.file_name}"
        await bot.download_file(file.file_path, file_path)
        
        file_data = parse_invoice_file(file_path)
        
        if not file_data:
            await message.answer("❌ Не удалось найти данные в файле. Проверьте формат.")
            return
        
        user_data_store[user_id].extend(file_data)
        
        file_stats = calculate_file_statistics(file_data)
        user_data = user_data_store[user_id]
        all_stats = calculate_statistics(user_data)
        
        response = f"""
📄 Файл обработан: {document.file_name}

Данные файла:
• Поездок в файле: {file_stats['trips_count']}
• Сумма в файле: {file_stats['total_amount']:,.0f} руб.

Общая статистика:
• Файлов загружено: {all_stats['unique_files']}
• Всего поездок: {all_stats['total_trips']}
• Автомобилей: {all_stats['unique_cars']}
• Водителей: {all_stats['unique_drivers']}
• Общая сумма: {all_stats['total_amount']:,.0f} руб.

💡 Используйте команды:
/report - полный отчет
/cars - по автомобилям  
/drivers - по водителям
        """
        
        await message.answer(response)
        
    except Exception as e:
        logger.error(f"Ошибка: {e}")
        await message.answer("❌ Произошла ошибка при обработке файла")

async def generate_report(message: Message, data, title):
    if not data:
        await message.answer("❌ Нет данных для отчета")
        return
    
    stats = calculate_statistics(data)
    
    car_reports = []
    for car_plate, car_data in stats['car_stats'].items():
        drivers = ', '.join(car_data['drivers'])
        files = ', '.join(list(car_data['files'])[:3])
        if len(car_data['files']) > 3:
            files += f" ... (еще {len(car_data['files']) - 3})"
        
        car_reports.append(f"🚗 {car_plate}\n"
                         f"• Поездок: {car_data['trips_count']}\n"
                         f"• Водители: {drivers}\n"
                         f"• Файлы: {files}\n"
                         f"• Общая сумма: {car_data['total_amount']:,.0f} руб.")
    
    driver_reports = []
    for driver, driver_data in stats['driver_stats'].items():
        if driver == "Фамилия не найдена":
            continue
        cars = ', '.join(driver_data['cars'])
        files = ', '.join(list(driver_data['files'])[:3])
        if len(driver_data['files']) > 3:
            files += f" ... (еще {len(driver_data['files']) - 3})"
        
        driver_reports.append(f"👤 {driver}\n"
                            f"• Поездок: {driver_data['trips_count']}\n"
                            f"• Автомобили: {cars}\n"
                            f"• Файлы: {files}\n"
                            f"• Общая сумма: {driver_data['total_amount']:,.0f} руб.")
    
    response = f"""
📊 {title}

Общая статистика:
• Файлов обработано: {stats['unique_files']}
• Всего поездок: {stats['total_trips']}
• Автомобилей: {stats['unique_cars']}  
• Водителей: {stats['unique_drivers']}
• Общая сумма: {stats['total_amount']:,.0f} руб.

По автомобилям:
{chr(10).join(car_reports)}

По водителям:
{chr(10).join(driver_reports)}

✅ Отчет сформирован!
    """
    
    if len(response) > 4000:
        parts = [response[i:i+4000] for i in range(0, len(response), 4000)]
        for part in parts:
            await message.answer(part)
            await asyncio.sleep(0.5)
    else:
        await message.answer(response)

@dp.message()
async def handle_text_message(message: Message):
    """Обработка текстовых сообщений для поиска по водителям и автомобилям"""
    user_id = message.from_user.id
    user_data = user_data_store[user_id]
    
    if not user_data:
        await message.answer("📭 Нет данных для поиска. Сначала отправьте файлы.")
        return
    
    search_text = message.text.strip()
    
    # Если это команда - пропускаем (обрабатывается другими хендлерами)
    if search_text.startswith('/'):
        return
    
    # Поиск по номеру автомобиля (цифры)
    if search_text.isdigit():
        car_results = [item for item in user_data if search_text in item['Гос_номер']]
        
        if car_results:
            car_plates = set(item['Гос_номер'] for item in car_results)
            total_trips = len(car_results)
            total_amount = sum(item['Стоимость'] for item in car_results)
            drivers = set(item['Водитель'] for item in car_results if item['Водитель'] != "Фамилия не найдена")
            files = set(item['Файл'] for item in car_results)
            
            response = f"""
🔍 РЕЗУЛЬТАТЫ ПОИСКА ПО АВТОМОБИЛЮ: {search_text}

Найдено автомобилей: {len(car_plates)}
• Номера: {', '.join(car_plates)}
• Поездок: {total_trips}
• Водители: {', '.join(drivers) if drivers else 'Не указаны'}
• Файлов: {len(files)}
• Общая сумма: {total_amount:,.0f} руб.

Детали поездок:
"""
            
            for i, item in enumerate(car_results[:10], 1):
                response += f"\n{i}. {item['Дата']} - {item['Водитель']} - {item['Стоимость']:,.0f} руб. ({item['Маршрут']})"
            
            if len(car_results) > 10:
                response += f"\n\n... и еще {len(car_results) - 10} поездок"
                
            await message.answer(response)
        else:
            await message.answer(f"❌ Автомобиль с номером '{search_text}' не найден")
    
    # Поиск по фамилии водителя (текст)
    else:
        driver_results = [item for item in user_data if search_text.lower() in item['Водитель'].lower()]
        
        if driver_results:
            drivers_found = set(item['Водитель'] for item in driver_results)
            total_trips = len(driver_results)
            total_amount = sum(item['Стоимость'] for item in driver_results)
            cars = set(item['Гос_номер'] for item in driver_results)
            files = set(item['Файл'] for item in driver_results)
            
            response = f"""
🔍 РЕЗУЛЬТАТЫ ПОИСКА ПО ВОДИТЕЛЮ: {search_text}

Найдено водителей: {len(drivers_found)}
• Фамилии: {', '.join(drivers_found)}
• Поездок: {total_trips}
• Автомобили: {', '.join(cars)}
• Файлов: {len(files)}
• Общая сумма: {total_amount:,.0f} руб.

Детали поездок:
"""
            
            for i, item in enumerate(driver_results[:10], 1):
                response += f"\n{i}. {item['Дата']} - {item['Гос_номер']} - {item['Стоимость']:,.0f} руб. ({item['Маршрут']})"
            
            if len(driver_results) > 10:
                response += f"\n\n... и еще {len(driver_results) - 10} поездок"
                
            await message.answer(response)
        else:
            await message.answer(f"❌ Водитель '{search_text}' не найден")

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
